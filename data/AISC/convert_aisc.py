"""Convert AISC shapes database Excel files to compact JSON for token-efficient lookup."""
import json
import openpyxl

KEEP = [
    "Type", "AISC_Manual_Label",
    "W", "A", "d", "tw", "bf", "tf", "kdes",
    # pipe / HSS round geometry
    "OD", "ID", "tnom", "tdes",
    # angle / double-angle geometry
    "b", "t",
    "x", "y",                      # centroid distances
    "Ix", "Zx", "Sx", "rx",
    "Iy", "Zy", "Sy", "ry",
    "Iz", "rz", "Sz",               # geometric (z) axis for single angles
    "J", "Cw",
    "bf/2tf", "h/tw", "b/t", "D/t",
]

def convert(xlsx_path, sheet_name, out_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    idx = {}
    for i, h in enumerate(headers):
        if h in KEEP and h not in idx:  # first occurrence only (imperial columns)
            idx[h] = i

    sections = []
    for row in rows:
        if not row[0]:
            continue
        record = {}
        for field in KEEP:
            if field not in idx:
                continue
            val = row[idx[field]]
            if isinstance(val, str) and val.strip() in ("", "—", "-"):
                val = None
            record[field] = val
        sections.append(record)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(sections, f, separators=(",", ":"))
    print(f"Wrote {len(sections)} sections to {out_path}")

KEEP_HIST = {
    # source column name  : output key
    "Edition":   "Edition",
    "Type":      "Type",
    "Designation": "AISC_Manual_Label",
    "A ":        "A",
    "d":         "d",
    "tw":        "tw",
    "bf":        "bf",
    "tf":        "tf",
    "k":         "kdes",
    "W":         "W",
    "Ix":        "Ix",
    "Zx":        "Zx",
    "Sx":        "Sx",
    "rx":        "rx",
    "Iy":        "Iy",
    "Zy":        "Zy",
    "Sy":        "Sy",
    "ry":        "ry",
    "bf/2tf":    "bf/2tf",
    "h/tw":      "h/tw",
}


def convert_historical(xlsx_path, sheet_name, out_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    raw_headers = list(next(rows))

    # Build index: source column name → column index (first occurrence)
    idx = {}
    for i, h in enumerate(raw_headers):
        if h in KEEP_HIST and h not in idx:
            idx[h] = i

    sections = []
    for row in rows:
        if not row[2]:  # Type column must be present
            continue
        record = {}
        for src, dst in KEEP_HIST.items():
            if src not in idx:
                continue
            val = row[idx[src]]
            if isinstance(val, str) and val.strip() in ("", "—", "-", "�"):
                val = None
            record[dst] = val
        # Only include rows that have a designation
        if record.get("AISC_Manual_Label"):
            sections.append(record)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(sections, f, separators=(",", ":"))
    print(f"Wrote {len(sections)} sections to {out_path}")


base = "d:/AI_TEST/Agent_Developer/Project_006_JuniorSE_Agent/data/AISC"

convert(
    f"{base}/aisc-shapes-database-v160-2.xlsx",
    "Database v16.0",
    f"{base}/sections_v16.json",
)

convert_historical(
    f"{base}/aisc-shapes-database-v160h.xlsx",
    "Database v16.0H",
    f"{base}/sections_v16h.json",
)
